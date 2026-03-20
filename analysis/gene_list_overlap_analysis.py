#!/usr/bin/env python3
"""
Gene list overlap analysis across three pigmentation gene sources.

Sources:
1. Raghunath et al. 2015 — melanocyte signaling network (node_properties from suppl.)
2. Baxter et al. 2018 — curated cross-species pigmentation gene list (Table S7)
3. Bajpai et al. 2023 — genome-wide CRISPR screen hits (Table S1, 10% FDR)

Goal: Assess realistic network size for grant proposal (claimed 400-500 nodes)
      and understand overlap between sources.
"""

import openpyxl
import os
import json
from collections import Counter

# ── Paths ───────────────────────────────────────────────────────────────────
DATA = os.path.join(os.path.dirname(__file__), '..', 'data')
OUTPUT = os.path.join(os.path.dirname(__file__), '..', 'output')
os.makedirs(OUTPUT, exist_ok=True)

# ═══════════════════════════════════════════════════════════════════════════
# 1. RAGHUNATH 2015 — Melanocyte signaling network
# ═══════════════════════════════════════════════════════════════════════════
print("=" * 70)
print("1. RAGHUNATH et al. 2015 — Melanocyte signaling network")
print("=" * 70)

wb_ragh = openpyxl.load_workbook(os.path.join(DATA, '13104_2015_1128_MOESM2_ESM.xlsx'))
ws_ragh = wb_ragh['node_properties']

# Extract all node names, strip _melan/_kerat suffixes
raw_nodes = set()
for row in ws_ragh.iter_rows(min_row=2, values_only=True):
    node = str(row[0]).strip() if row[0] else ''
    if node:
        for suffix in ['_melan', '_kerat']:
            if node.endswith(suffix):
                node = node[:-len(suffix)]
        raw_nodes.add(node)

# Filter out non-gene entries (metabolites, processes, complexes)
NON_GENE_ENTRIES = {
    # Small molecules / metabolites
    '4HNE', 'ACTH', 'Arachidonic_Acid', 'cAMP', 'DAG', 'IP3', 'NO',
    'PGE2', 'PGF2a', 'ROS', 'Ca2+', 'PKC', 'Melanin', 'cGMP',
    'Calcium_cyt', 'Ceramide', 'Cysteine', 'Cysteinyl_DOPA',
    'DHI', 'DHICA', 'DOPA', 'DOPAchrome', 'DOPAquinone',
    'Eumelanin', 'GSH', 'Glutathionyl_DOPA', 'Pheomelanin',
    'Indole_5_6_quinone', 'Indole_5_6_quinone_carboxylic_acid',
    'Tyrosine', 'alpha_MSH', 'Nitric_oxide', 'Singlet_oxygen',
    'Trypsin', 'FICZ',
    # Processes / phenotypes
    'Apoptosis', 'Cell_cycle_arrest', 'Cell_differentiation',
    'Cell_proliferation', 'Cell_survival', 'DNA_Damage', 'DNA_Repair',
    'Dendrite_formation', 'Lipid_Peroxidation', 'Melanocyte_migration',
    'Melanosome_biogenesis', 'Melanosome_phagocytosis',
    'Skin_aging', 'Skin_inflammation',
    # UV / stimuli
    'UVR', 'UVA', 'UVB',
    # Generic labels
    'Node', 'MMPs', 'PLA2', 'PLC',
    # Odd entry
    'PhosphodiesteHRASe',
}

raghunath_genes = set()
raghunath_non_genes = set()
for node in raw_nodes:
    if ':' in node:  # complexes like AHR:SRC:HSP90AA1
        raghunath_non_genes.add(node)
    elif node in NON_GENE_ENTRIES:
        raghunath_non_genes.add(node)
    else:
        raghunath_genes.add(node)

# Standardize: ensure all uppercase (some like F10, F7 are fine)
raghunath_genes = {g.upper() for g in raghunath_genes}

print(f"  Raw unique nodes (after collapsing _melan/_kerat): {len(raw_nodes)}")
print(f"  Non-gene entries removed: {len(raghunath_non_genes)}")
print(f"  Clean gene symbols: {len(raghunath_genes)}")
print()

# ═══════════════════════════════════════════════════════════════════════════
# 2. BAXTER 2018 — Curated pigmentation gene list
# ═══════════════════════════════════════════════════════════════════════════
print("=" * 70)
print("2. BAXTER et al. 2018 — Curated pigmentation gene list (Table S7)")
print("=" * 70)

wb_bax = openpyxl.load_workbook(
    os.path.join(DATA, 'baxter2018_650_pigmentation_genes_tableS7.xlsx'))
ws_bax = wb_bax['650 Pigmentation Genes']

baxter_all = set()
baxter_human_phenotype = set()
baxter_by_species = {}

for row in ws_bax.iter_rows(min_row=2, max_col=12, values_only=True):
    symbol = row[1]  # Human gene symbol (col B)
    species = str(row[11]).strip() if row[11] else ''
    if symbol and isinstance(symbol, str) and not symbol.startswith('ENSG') \
       and not symbol.startswith('ENSDARG') and len(symbol) < 20:
        sym = symbol.strip().upper()
        baxter_all.add(sym)
        if species:
            if species not in baxter_by_species:
                baxter_by_species[species] = set()
            baxter_by_species[species].add(sym)
        # Genes with human phenotype evidence
        if 'Human' in species or 'human' in species:
            baxter_human_phenotype.add(sym)

print(f"  Total unique human gene symbols: {len(baxter_all)}")
print(f"  Genes with human phenotype evidence: {len(baxter_human_phenotype)}")
print(f"  By species phenotype breakdown:")
for sp in sorted(baxter_by_species.keys()):
    print(f"    {sp}: {len(baxter_by_species[sp])}")
print()

# ═══════════════════════════════════════════════════════════════════════════
# 3. BAJPAI 2023 — CRISPR screen hits
# ═══════════════════════════════════════════════════════════════════════════
print("=" * 70)
print("3. BAJPAI et al. 2023 — CRISPR screen (Table S1)")
print("=" * 70)

wb_baj = openpyxl.load_workbook(
    os.path.join(DATA, 'bajpai2023_crispr_screen_tableS1.xlsx'))
ws_baj = wb_baj['Low SSC FACS enriched genes']

bajpai_all_screened = set()
bajpai_promoting = set()  # melanin-promoting at 10% FDR
bajpai_promoting_5pct = set()

for row in ws_baj.iter_rows(min_row=2, values_only=True):
    symbol = str(row[1]).strip().upper() if row[1] else ''
    effect = row[12]  # Combined_casTLE_Effect
    qval = row[17]    # q_value
    if symbol:
        bajpai_all_screened.add(symbol)
        if qval is not None and effect is not None:
            if qval <= 0.10 and effect > 0:
                bajpai_promoting.add(symbol)
            if qval <= 0.05 and effect > 0:
                bajpai_promoting_5pct.add(symbol)

print(f"  Total genes screened: {len(bajpai_all_screened)}")
print(f"  Melanin-promoting hits (10% FDR): {len(bajpai_promoting)}")
print(f"  Melanin-promoting hits (5% FDR): {len(bajpai_promoting_5pct)}")
print()

# ═══════════════════════════════════════════════════════════════════════════
# 4. OVERLAP ANALYSIS
# ═══════════════════════════════════════════════════════════════════════════
print("=" * 70)
print("4. OVERLAP ANALYSIS")
print("=" * 70)

# Pairwise overlaps
ragh_bax = raghunath_genes & baxter_all
ragh_baj = raghunath_genes & bajpai_promoting
bax_baj = baxter_all & bajpai_promoting

# Triple overlap
all_three = raghunath_genes & baxter_all & bajpai_promoting

# Union
union_all = raghunath_genes | baxter_all | bajpai_promoting

# Only in one source
only_ragh = raghunath_genes - baxter_all - bajpai_promoting
only_bax = baxter_all - raghunath_genes - bajpai_promoting
only_baj = bajpai_promoting - raghunath_genes - baxter_all

# In exactly 2
ragh_bax_only = (ragh_bax) - bajpai_promoting
ragh_baj_only = (ragh_baj) - baxter_all
bax_baj_only = (bax_baj) - raghunath_genes

print(f"\n  SET SIZES:")
print(f"    Raghunath (network nodes):     {len(raghunath_genes):>4}")
print(f"    Baxter (curated list):         {len(baxter_all):>4}")
print(f"    Bajpai (CRISPR 10% FDR):       {len(bajpai_promoting):>4}")
print(f"    UNION of all three:            {len(union_all):>4}")

print(f"\n  PAIRWISE OVERLAPS:")
print(f"    Raghunath ∩ Baxter:            {len(ragh_bax):>4}  ({len(ragh_bax)/len(raghunath_genes)*100:.1f}% of Raghunath)")
print(f"    Raghunath ∩ Bajpai:            {len(ragh_baj):>4}  ({len(ragh_baj)/len(raghunath_genes)*100:.1f}% of Raghunath)")
print(f"    Baxter ∩ Bajpai:               {len(bax_baj):>4}  ({len(bax_baj)/len(bajpai_promoting)*100:.1f}% of Bajpai)")

print(f"\n  THREE-WAY OVERLAP:")
print(f"    All three sources:             {len(all_three):>4}")

print(f"\n  EXCLUSIVE TO ONE SOURCE:")
print(f"    Only Raghunath:                {len(only_ragh):>4}")
print(f"    Only Baxter:                   {len(only_bax):>4}")
print(f"    Only Bajpai:                   {len(only_baj):>4}")

print(f"\n  IN EXACTLY TWO SOURCES:")
print(f"    Raghunath + Baxter only:       {len(ragh_bax_only):>4}")
print(f"    Raghunath + Bajpai only:       {len(ragh_baj_only):>4}")
print(f"    Baxter + Bajpai only:          {len(bax_baj_only):>4}")

# ── Venn diagram verification ──
venn_total = (len(only_ragh) + len(only_bax) + len(only_baj) +
              len(ragh_bax_only) + len(ragh_baj_only) + len(bax_baj_only) +
              len(all_three))
print(f"\n  Venn sum check: {venn_total} (should equal union: {len(union_all)})")

# ═══════════════════════════════════════════════════════════════════════════
# 5. GRANT NETWORK SIZE ASSESSMENT
# ═══════════════════════════════════════════════════════════════════════════
print("\n" + "=" * 70)
print("5. GRANT NETWORK SIZE ASSESSMENT")
print("=" * 70)

print(f"""
  The grant claims a network of '400-500 nodes'.

  REALITY CHECK:
  • Raghunath network has {len(raghunath_genes)} unique gene nodes
    (plus {len(raghunath_non_genes)} non-gene nodes like metabolites/processes)
  • Baxter curated list has {len(baxter_all)} genes (cross-species)
    - But only {len(baxter_human_phenotype)} have human phenotype evidence
  • Bajpai CRISPR screen identified {len(bajpai_promoting)} melanin-promoting genes

  UNION of all three sources: {len(union_all)} unique genes

  If we BUILD A NETWORK using all three sources:
  • Start with Raghunath network structure ({len(raghunath_genes)} genes)
  • Add Baxter genes not in Raghunath: +{len(baxter_all - raghunath_genes)} genes
  • Add Bajpai genes not in either: +{len(bajpai_promoting - raghunath_genes - baxter_all)} genes
  • Total potential nodes: {len(union_all)}

  The 400-500 estimate is {"PLAUSIBLE" if 400 <= len(union_all) <= 600 else "TOO LOW" if len(union_all) > 600 else "TOO HIGH if only using Raghunath"}.

  However, a meaningful NETWORK requires edges (interactions), not just
  a gene list. The Raghunath network only has edges for {len(raghunath_genes)} genes.
  Adding Baxter/Bajpai genes would require querying STRING/BioGRID for
  interactions, which would create a much larger but sparser network.
""")

# ═══════════════════════════════════════════════════════════════════════════
# 6. DETAILED GENE LISTS
# ═══════════════════════════════════════════════════════════════════════════
print("=" * 70)
print("6. KEY GENE LISTS")
print("=" * 70)

print(f"\n  ALL THREE SOURCES ({len(all_three)} genes):")
for g in sorted(all_three):
    print(f"    {g}")

print(f"\n  BAJPAI CRISPR HITS *NOT* IN RAGHUNATH OR BAXTER ({len(only_baj)} truly novel):")
for g in sorted(only_baj):
    print(f"    {g}")

print(f"\n  RAGHUNATH NETWORK GENES ALSO IN BAXTER ({len(ragh_bax)} genes):")
for g in sorted(ragh_bax):
    print(f"    {g}")

# ═══════════════════════════════════════════════════════════════════════════
# 7. SAVE RESULTS
# ═══════════════════════════════════════════════════════════════════════════
results = {
    'raghunath_genes': sorted(raghunath_genes),
    'raghunath_non_genes': sorted(raghunath_non_genes),
    'baxter_all': sorted(baxter_all),
    'baxter_human_phenotype': sorted(baxter_human_phenotype),
    'bajpai_promoting_10pct_fdr': sorted(bajpai_promoting),
    'bajpai_promoting_5pct_fdr': sorted(bajpai_promoting_5pct),
    'overlap_all_three': sorted(all_three),
    'overlap_raghunath_baxter': sorted(ragh_bax),
    'overlap_raghunath_bajpai': sorted(ragh_baj),
    'overlap_baxter_bajpai': sorted(bax_baj),
    'only_raghunath': sorted(only_ragh),
    'only_baxter': sorted(only_bax),
    'only_bajpai': sorted(only_baj),
    'union_all': sorted(union_all),
    'summary': {
        'raghunath_n': len(raghunath_genes),
        'baxter_n': len(baxter_all),
        'bajpai_n': len(bajpai_promoting),
        'union_n': len(union_all),
        'all_three_n': len(all_three),
        'ragh_bax_n': len(ragh_bax),
        'ragh_baj_n': len(ragh_baj),
        'bax_baj_n': len(bax_baj),
    }
}

outpath = os.path.join(OUTPUT, 'gene_list_overlap_results.json')
with open(outpath, 'w') as f:
    json.dump(results, f, indent=2)
print(f"\n  Results saved to: {outpath}")

# Also save a simple CSV of the union with source annotations
csv_path = os.path.join(OUTPUT, 'gene_list_union_annotated.csv')
with open(csv_path, 'w') as f:
    f.write('gene,in_raghunath,in_baxter,in_bajpai_10pct,source_count\n')
    for g in sorted(union_all):
        r = 1 if g in raghunath_genes else 0
        b = 1 if g in baxter_all else 0
        j = 1 if g in bajpai_promoting else 0
        f.write(f'{g},{r},{b},{j},{r+b+j}\n')
print(f"  Annotated union CSV saved to: {csv_path}")

print("\n✓ Analysis complete.")
