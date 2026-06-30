"""
pool_venn_gene_lists.py

Produces a single pooled gene table for all four studies represented in
figure_gene_source_venn.py (Raghunath 2015, Baxter 2018, GWAS Catalog,
Bajpai 2023 CRISPR screen) with per-study membership flags and GRCh38
gene coordinates from mygene.info.

Outputs:
  output/pooled_venn_genes.tsv  — one row per unique gene symbol

Columns:
  gene            — HGNC symbol (upper-cased)
  in_raghunath    — 1/0
  in_baxter       — 1/0
  in_gwas         — 1/0 (≥2 independent associations)
  in_bajpai       — 1/0 (q ≤ 0.10, effect > 0)
  n_studies       — count of sets this gene appears in
  chrom           — GRCh38 chromosome (mygene.info)
  start           — GRCh38 start (1-based, inclusive)
  end             — GRCh38 end   (1-based, inclusive)
  strand          — + / -
  ensembl_id      — Ensembl gene ID (first hit)
  coord_source    — "gene_regions_bed" | "mygene.info" | "not_found"
  reference_genome — "GRCh38" where coords were found, else ""

Usage:
  pip install openpyxl pandas requests
  python analysis/pool_venn_gene_lists.py [--base /path/to/repo]

Gene parsing is identical to figure_gene_source_venn.py so membership
flags are consistent with the Venn figure.
"""

import argparse
import os
import sys
import time

import openpyxl
import pandas as pd
import requests

# ── Argument parsing ──────────────────────────────────────────────────────────
ap = argparse.ArgumentParser(description=__doc__,
                             formatter_class=argparse.RawDescriptionHelpFormatter)
ap.add_argument("--base",
                default=os.path.dirname(os.path.dirname(os.path.abspath(__file__))),
                help="Project root (default: repo root)")
args = ap.parse_args()

DATA = os.path.join(args.base, "data")
OUT  = os.path.join(args.base, "output")
os.makedirs(OUT, exist_ok=True)

# ── 1. Raghunath 2015 ─────────────────────────────────────────────────────────
# Parsing identical to figure_gene_source_venn.py:27-50
NON_GENE = {
    '4HNE','ACTH','Arachidonic_Acid','cAMP','DAG','IP3','NO','PGE2','PGF2a',
    'ROS','Ca2+','PKC','Melanin','cGMP','Calcium_cyt','Ceramide','Cysteine',
    'Cysteinyl_DOPA','DHI','DHICA','DOPA','DOPAchrome','DOPAquinone',
    'Eumelanin','GSH','Glutathionyl_DOPA','Pheomelanin','Indole_5_6_quinone',
    'Indole_5_6_quinone_carboxylic_acid','Tyrosine','alpha_MSH','Nitric_oxide',
    'Singlet_oxygen','Trypsin','FICZ','Apoptosis','Cell_cycle_arrest',
    'Cell_differentiation','Cell_proliferation','Cell_survival','DNA_Damage',
    'DNA_Repair','Dendrite_formation','Lipid_Peroxidation',
    'Melanocyte_migration','Melanosome_biogenesis','Melanosome_phagocytosis',
    'Skin_aging','Skin_inflammation','UVR','UVA','UVB','Node','MMPs','PLA2',
    'PLC','PhosphodiesteHRASe',
}
wb = openpyxl.load_workbook(os.path.join(DATA, '13104_2015_1128_MOESM2_ESM.xlsx'))
ws = wb['node_properties']
raghunath = set()
for row in ws.iter_rows(min_row=2, values_only=True):
    n = str(row[0]).strip() if row[0] else ''
    if not n: continue
    for suf in ('_melan', '_kerat'):
        if n.endswith(suf): n = n[:-len(suf)]
    if ':' in n or n in NON_GENE: continue
    raghunath.add(n.upper())
print(f"Raghunath 2015:   {len(raghunath):>4d} genes")

# ── 2. Baxter 2018 ────────────────────────────────────────────────────────────
# Parsing identical to figure_gene_source_venn.py:52-62
wb = openpyxl.load_workbook(
    os.path.join(DATA, 'baxter2018_650_pigmentation_genes_tableS7.xlsx'))
ws = wb['650 Pigmentation Genes']
baxter = set()
for row in ws.iter_rows(min_row=2, max_col=12, values_only=True):
    sym = row[1]
    if (sym and isinstance(sym, str)
            and not sym.startswith('ENSG') and not sym.startswith('ENSDARG')
            and len(sym) < 20):
        baxter.add(sym.strip().upper())
print(f"Baxter 2018:      {len(baxter):>4d} genes")

# ── 3. GWAS Catalog ───────────────────────────────────────────────────────────
# Parsing identical to figure_gene_source_venn.py:64-68
gwas_df = pd.read_csv(os.path.join(DATA, 'gwas_pigmentation_associations.csv'))
gwas_df['gene'] = gwas_df['gene'].astype(str).str.upper()
gwas_counts = gwas_df.groupby('gene').size()
gwas = set(gwas_counts[gwas_counts >= 2].index)
print(f"GWAS Catalog:     {len(gwas):>4d} genes (≥2 associations)")

# ── 4. Bajpai 2023 ────────────────────────────────────────────────────────────
# Parsing identical to figure_gene_source_venn.py:70-79
wb = openpyxl.load_workbook(
    os.path.join(DATA, 'bajpai2023_crispr_screen_tableS1.xlsx'))
ws = wb['Low SSC FACS enriched genes']
bajpai = set()
for row in ws.iter_rows(min_row=2, values_only=True):
    sym = str(row[1]).strip().upper() if row[1] else ''
    eff = row[12]; q = row[17]
    if sym and q is not None and eff is not None and q <= 0.10 and eff > 0:
        bajpai.add(sym)
print(f"Bajpai 2023:      {len(bajpai):>4d} genes (q≤0.10, effect>0)")

# ── 5. Pool all genes ─────────────────────────────────────────────────────────
all_genes = sorted(raghunath | baxter | gwas | bajpai)
print(f"\nUnion (all studies): {len(all_genes)} unique genes")

df = pd.DataFrame({"gene": all_genes})
df["in_raghunath"] = df["gene"].isin(raghunath).astype(int)
df["in_baxter"]    = df["gene"].isin(baxter).astype(int)
df["in_gwas"]      = df["gene"].isin(gwas).astype(int)
df["in_bajpai"]    = df["gene"].isin(bajpai).astype(int)
df["n_studies"]    = (df["in_raghunath"] + df["in_baxter"]
                      + df["in_gwas"]    + df["in_bajpai"])

# ── 6. Seed coordinates from existing gene_regions.bed (Raghunath 128 genes) ─
# These are already GRCh38 (gnomAD pipeline reference).
coord_cols = ["chrom", "start", "end", "strand", "ensembl_id",
              "coord_source", "reference_genome"]
for col in coord_cols:
    df[col] = ""

bed_path = os.path.join(DATA, "gene_regions.bed")
if os.path.exists(bed_path):
    bed = pd.read_csv(bed_path, sep="\t", header=None,
                      names=["chrom", "start", "end", "gene"])
    bed["gene"] = bed["gene"].str.upper()
    bed_idx = bed.set_index("gene")
    mask = df["gene"].isin(bed_idx.index)
    for col in ("chrom", "start", "end"):
        df.loc[mask, col] = df.loc[mask, "gene"].map(bed_idx[col]).astype(str)
    df.loc[mask, "coord_source"]    = "gene_regions_bed"
    df.loc[mask, "reference_genome"] = "GRCh38"
    print(f"Seeded {mask.sum()} genes from gene_regions.bed (GRCh38)")

# ── 7. Fill remaining via mygene.info (GRCh38) ───────────────────────────────
# Batch POST endpoint — free, no auth, 1000 genes per request.
MYGENE_URL  = "https://mygene.info/v3/gene"
MYGENE_POST = "https://mygene.info/v3/query"
BATCH       = 500
SLEEP       = 0.25

still_needed = df.loc[df["coord_source"] == "", "gene"].tolist()
print(f"Querying mygene.info for {len(still_needed)} remaining genes...")

def batch_query(symbols):
    """POST a batch of symbols; return list of hit dicts."""
    try:
        r = requests.post(
            MYGENE_POST,
            data={
                "q":      ",".join(symbols),
                "scopes": "symbol,alias",
                "fields": "symbol,genomic_pos,ensembl.gene",
                "species": "human",
                "size":   len(symbols),
            },
            timeout=60,
        )
        r.raise_for_status()
        return r.json()
    except requests.RequestException as e:
        print(f"  ! mygene.info request failed: {e}", file=sys.stderr)
        return []

results = {}
for i in range(0, len(still_needed), BATCH):
    chunk = still_needed[i:i + BATCH]
    hits  = batch_query(chunk)
    time.sleep(SLEEP)
    for hit in hits:
        if hit.get("notfound") or "genomic_pos" not in hit:
            continue
        # mygene.info may return a list when there are multiple loci
        sym = hit.get("symbol", "").upper()
        if not sym:
            continue
        gpos = hit["genomic_pos"]
        if isinstance(gpos, list):
            # pick canonical chromosome (not patch/alt)
            gpos = next((g for g in gpos if not str(g.get("chr","")).startswith("H")), gpos[0])
        chrom  = str(gpos.get("chr", ""))
        start  = gpos.get("start", "")
        end    = gpos.get("end", "")
        strand = "+" if gpos.get("strand", 1) == 1 else "-"
        ensembl = hit.get("ensembl", {})
        if isinstance(ensembl, list):
            ensembl = ensembl[0]
        ens_id = ensembl.get("gene", "") if isinstance(ensembl, dict) else ""
        results[sym] = (chrom, start, end, strand, ens_id)

n_found = 0
for sym, (chrom, start, end, strand, ens_id) in results.items():
    idx = df.index[df["gene"] == sym]
    if len(idx) == 0:
        continue
    df.loc[idx, "chrom"]            = chrom
    df.loc[idx, "start"]            = start
    df.loc[idx, "end"]              = end
    df.loc[idx, "strand"]           = strand
    df.loc[idx, "ensembl_id"]       = ens_id
    df.loc[idx, "coord_source"]     = "mygene.info"
    df.loc[idx, "reference_genome"] = "GRCh38"
    n_found += 1

n_missing = (df["coord_source"] == "").sum()
print(f"  mygene.info resolved:  {n_found}")
print(f"  not found / no coords: {n_missing}")
df.loc[df["coord_source"] == "", "coord_source"] = "not_found"

# ── 8. Sort and write ─────────────────────────────────────────────────────────
# Coerce numeric cols for sort; non-numeric (patch chroms) sort last
df["_chrom_sort"] = pd.to_numeric(
    df["chrom"].str.replace("chr", "", regex=False).str.replace("X","23").str.replace("Y","24"),
    errors="coerce")
df["_pos_sort"]   = pd.to_numeric(df["start"], errors="coerce")
df = (df.sort_values(["_chrom_sort", "_pos_sort", "gene"],
                     na_position="last")
        .drop(columns=["_chrom_sort", "_pos_sort"])
        .reset_index(drop=True))

out_path = os.path.join(OUT, "pooled_venn_genes.tsv")
df.to_csv(out_path, sep="\t", index=False)
print(f"\nWrote {len(df)} genes → {out_path}")

# ── 9. Summary ────────────────────────────────────────────────────────────────
print("\nStudy membership breakdown:")
for n in range(4, 0, -1):
    sub = df[df["n_studies"] == n]
    label = {4:"all 4", 3:"any 3", 2:"any 2", 1:"only 1"}[n]
    print(f"  {label}: {len(sub):>4d} genes")
print(f"\nCoordinate coverage:")
for src in ("gene_regions_bed", "mygene.info", "not_found"):
    print(f"  {src}: {(df['coord_source']==src).sum()}")
if n_missing:
    print("\nGenes with no coordinates found:")
    print("  " + ", ".join(df.loc[df["coord_source"]=="not_found","gene"].tolist()))
