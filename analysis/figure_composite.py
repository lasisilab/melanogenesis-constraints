"""
figure_composite.py

Three-panel composite figure:
  A. Tissue specificity (τ) vs. LoF intolerance — gene source highlights
  B. MC1R nucleotide diversity vs. dataset median across populations
  C. PBS quadrant scatter — African vs. Melanesian branch selection signals

Layout:
  Left column:  A (top, scatter) + B (bottom, bars)
  Right column: C (full height, square scatter)

Output: output/figure_composite.{png,pdf}
"""
import os
import openpyxl
import numpy as np
import pandas as pd
import matplotlib
matplotlib.use('Agg')
import matplotlib.pyplot as plt
import matplotlib.lines as mlines
from scipy import stats
from adjustText import adjust_text

PROJECT_DIR = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
DATA_DIR    = os.path.join(PROJECT_DIR, 'data')
OUT_DIR     = os.path.join(PROJECT_DIR, 'output')

# ══════════════════════════════════════════════════════════════════════════════
# DATA LOADING
# ══════════════════════════════════════════════════════════════════════════════
master = pd.read_csv(os.path.join(DATA_DIR, 'network_constraint_categorized.csv'))
master['gene'] = master['gene'].str.upper()

# Baxter 2018
wb = openpyxl.load_workbook(
    os.path.join(DATA_DIR, 'baxter2018_650_pigmentation_genes_tableS7.xlsx'))
ws = wb['650 Pigmentation Genes']
baxter = set()
for row in ws.iter_rows(min_row=2, max_col=12, values_only=True):
    sym = row[1]
    if (sym and isinstance(sym, str)
            and not sym.startswith('ENSG') and not sym.startswith('ENSDARG')
            and len(sym) < 20):
        baxter.add(sym.strip().upper())

# Bajpai 2023
wb = openpyxl.load_workbook(
    os.path.join(DATA_DIR, 'bajpai2023_crispr_screen_tableS1.xlsx'))
ws = wb['Low SSC FACS enriched genes']
bajpai = set()
for row in ws.iter_rows(min_row=2, values_only=True):
    sym = str(row[1]).strip().upper() if row[1] else ''
    eff = row[12]; q = row[17]
    if sym and q is not None and eff is not None and q <= 0.10 and eff > 0:
        bajpai.add(sym)

# GWAS Catalog (≥1 association)
gwas_df = pd.read_csv(os.path.join(DATA_DIR, 'gwas_pigmentation_associations.csv'))
gwas = set(gwas_df['gene'].astype(str).str.upper().unique())

# Pi (nucleotide diversity)
pi_df = pd.read_csv(os.path.join(DATA_DIR, 'pi_per_gene.csv'))
pi_df['gene'] = pi_df['gene'].str.upper()

# PBS
pbs = pd.read_csv(os.path.join(DATA_DIR, 'pbs_per_gene.csv'))
pbs['gene'] = pbs['gene'].str.upper()
pbs['pbs1_african']    = pbs['pbs1_african'].clip(lower=0)
pbs['pbs3_melanesian'] = pbs['pbs3_melanesian'].clip(lower=0)

# ══════════════════════════════════════════════════════════════════════════════
# PANEL A — Tau vs LOEUF with gene-source highlights
# ══════════════════════════════════════════════════════════════════════════════
df_A = master[['gene', 'tau', 'LOEUF']].dropna(subset=['tau', 'LOEUF']).copy()
df_A['in_gwas']   = df_A['gene'].isin(gwas)
df_A['in_baxter'] = df_A['gene'].isin(baxter)
df_A['in_bajpai'] = df_A['gene'].isin(bajpai)

def assign_combo(r):
    g, b, p = r['in_gwas'], r['in_baxter'], r['in_bajpai']
    if g and b and p: return 'All three'
    if g and b:       return 'GWAS + Baxter'
    if g:             return 'GWAS only'
    if b:             return 'Baxter only'
    if p:             return 'Bajpai only'
    return 'None'

df_A['combo'] = df_A.apply(assign_combo, axis=1)

COMBO_ORDER = ['None', 'Baxter only', 'GWAS only', 'Bajpai only',
               'GWAS + Baxter', 'All three']
COMBO_STYLE = {
    'None':          dict(color='#CFCFCF', marker='o', s=50,  alpha=0.70, zorder=2, lw=0.4),
    'Baxter only':   dict(color='#3D77C9', marker='s', s=140, alpha=0.82, zorder=3, lw=0.6),
    'GWAS only':     dict(color='#E8A33A', marker='o', s=150, alpha=0.88, zorder=4, lw=0.6),
    'Bajpai only':   dict(color='#5DA85D', marker='D', s=150, alpha=0.90, zorder=4, lw=0.6),
    'GWAS + Baxter': dict(color='#9B59B6', marker='P', s=200, alpha=0.90, zorder=5, lw=0.6),
    'All three':     dict(color='#C0392B', marker='*', s=310, alpha=0.95, zorder=6, lw=0.6),
}
COMBO_LABEL = {
    'None':          'Not in any external source',
    'Baxter only':   'Baxter 2018 only',
    'GWAS only':     'GWAS Catalog ≥1 only',
    'Bajpai only':   'Bajpai 2023 CRISPR only',
    'GWAS + Baxter': 'GWAS + Baxter',
    'All three':     'GWAS + Baxter + Bajpai',
}
LABEL_GENES_A = {
    'TYR', 'TYRP1', 'DCT', 'PMEL', 'MC1R', 'OCA2', 'MLANA',
    'MITF', 'SOX10', 'PAX3', 'TFAP2A', 'KIT', 'KITLG', 'EDNRB',
    'CCND1', 'EZR', 'IL6', 'MAP3K1', 'SPTLC2',
}
rho_A, p_A = stats.spearmanr(df_A['tau'], df_A['LOEUF'])
print(f"Panel A: Spearman ρ={rho_A:.3f}, p={p_A:.2e}, n={len(df_A)}")

# ══════════════════════════════════════════════════════════════════════════════
# PANEL B — MC1R pi bars
# ══════════════════════════════════════════════════════════════════════════════
POPS       = ['african', 'melanesian', 'southasian', 'european', 'eastasian']
POP_LABELS = {'african': 'African', 'melanesian': 'Melanesian',
              'southasian': 'South Asian', 'european': 'European',
              'eastasian': 'East Asian'}
POP_COLORS = {'african': '#C0392B', 'melanesian': '#2471A3',
              'southasian': '#1E8449', 'european': '#7D3C98',
              'eastasian': '#D4A017'}
mc1r_row  = pi_df[pi_df['gene'] == 'MC1R'].iloc[0]
mc1r_vals = [mc1r_row[f'pi_{p}'] for p in POPS]
med_vals  = [pi_df[f'pi_{p}'].median()  for p in POPS]

# ══════════════════════════════════════════════════════════════════════════════
# PANEL C — PBS quadrants
# ══════════════════════════════════════════════════════════════════════════════
df_C = master.merge(pbs[['gene', 'pbs1_african', 'pbs3_melanesian']],
                    on='gene', how='left')
df_C = df_C.dropna(subset=['pbs1_african', 'pbs3_melanesian'])

thr_a = df_C['pbs1_african'].quantile(0.75)
thr_m = df_C['pbs3_melanesian'].quantile(0.75)

def assign_quad(row):
    hi_a = row['pbs1_african']    >= thr_a
    hi_m = row['pbs3_melanesian'] >= thr_m
    if hi_a and hi_m: return 'Both'
    if hi_a:           return 'AfricanOnly'
    if hi_m:           return 'MelanesianOnly'
    return 'Neither'

df_C['quadrant'] = df_C.apply(assign_quad, axis=1)

COLOR_C = {'AfricanOnly':    '#6600cc',
           'MelanesianOnly': '#9ECC1B',
           'Both':           '#00C09A',
           'Neither':        '#CFCFCF'}
LABEL_C = {
    'AfricanOnly':    'African-specific selection',
    'MelanesianOnly': 'Melanesian-specific selection',
    'Both':           'Shared selection (both populations)',
    'Neither':        'No selection signal',
}
ALWAYS_LABEL_C = {
    'TYR', 'TYRP1', 'DCT', 'PMEL', 'MC1R', 'OCA2', 'MLANA',
    'MITF', 'SOX10', 'PAX3', 'TFAP2A', 'KIT', 'KITLG', 'EDNRB',
}
df_C['rank_sum'] = df_C['pbs1_african'].rank() + df_C['pbs3_melanesian'].rank()
labels_C = ALWAYS_LABEL_C | set(df_C.nlargest(15, 'rank_sum')['gene'])
print(f"Panel C: {len(df_C)} genes, thresholds African={thr_a:.3f} Melanesian={thr_m:.3f}")

# ══════════════════════════════════════════════════════════════════════════════
# BUILD FIGURE
# ══════════════════════════════════════════════════════════════════════════════
fig = plt.figure(figsize=(20, 18))
gs = fig.add_gridspec(2, 2,
                      height_ratios=[1.55, 1.0],
                      width_ratios=[1.0, 1.0])
ax_A = fig.add_subplot(gs[0, 0])
ax_B = fig.add_subplot(gs[1, 0])
ax_C = fig.add_subplot(gs[:, 1])

# ── Panel A ──────────────────────────────────────────────────────────────────
handles_A = []
for combo in COMBO_ORDER:
    sub = df_A[df_A['combo'] == combo]
    if len(sub) == 0:
        continue
    st = COMBO_STYLE[combo]
    ax_A.scatter(sub['tau'], sub['LOEUF'],
                 c=st['color'], marker=st['marker'], s=st['s'],
                 alpha=st['alpha'], edgecolors='white', linewidths=st['lw'],
                 zorder=st['zorder'])
    handles_A.append(mlines.Line2D(
        [], [], color=st['color'], marker=st['marker'],
        markersize=8, linestyle='None',
        markeredgecolor='white', markeredgewidth=0.5,
        label=f"{COMBO_LABEL[combo]}  (n={len(sub)})"))

x_line = np.linspace(df_A['tau'].min(), df_A['tau'].max(), 200)
sl, ic, *_ = stats.linregress(df_A['tau'], df_A['LOEUF'])
ax_A.plot(x_line, sl * x_line + ic, color='#888', lw=1.0, ls='--', zorder=1)

ax_A.text(0.03, 0.97,
          f'Spearman ρ = {rho_A:.3f},  p = {p_A:.2e}\nn = {len(df_A)} genes',
          transform=ax_A.transAxes, fontsize=9, va='top',
          bbox=dict(boxstyle='round,pad=0.4', facecolor='white',
                    edgecolor='#aaa', alpha=0.92))

texts_A = []
for _, r in df_A.iterrows():
    if r['gene'] in LABEL_GENES_A:
        texts_A.append(ax_A.text(r['tau'], r['LOEUF'], r['gene'],
                                  fontsize=7.5, fontweight='bold', style='italic',
                                  color='#111', zorder=7))
adjust_text(texts_A, ax=ax_A,
            arrowprops=dict(arrowstyle='-', color='#aaa', lw=0.5),
            force_points=(1.2, 1.2), force_text=(1.2, 1.2),
            expand_points=(1.8, 1.8), expand_text=(1.5, 1.5))

ax_A.set_xlabel('Tissue specificity (τ)\n← uniform · · · tissue-specific →',
                fontsize=11, labelpad=6)
ax_A.set_ylabel('LOEUF  (higher = less constrained)', fontsize=11)
ax_A.set_title('Tissue specificity vs. LoF intolerance — gene source highlights',
               fontsize=11.5, fontweight='bold', loc='left', pad=8)
ax_A.legend(handles=handles_A,
            loc='upper center', bbox_to_anchor=(0.5, -0.20), ncol=2,
            fontsize=8.5, framealpha=0.92, edgecolor='#999',
            title='Gene source combination', title_fontsize=9)
ax_A.tick_params(labelsize=10)
ax_A.spines['top'].set_visible(False)
ax_A.spines['right'].set_visible(False)
ax_A.text(-0.10, 1.06, 'A', transform=ax_A.transAxes,
          fontsize=20, fontweight='bold', va='top')

# ── Panel B ──────────────────────────────────────────────────────────────────
x = np.arange(len(POPS))
w = 0.35
ax_B.bar(x - w/2, mc1r_vals, width=w, zorder=3,
         color=[POP_COLORS[p] for p in POPS], alpha=0.90,
         edgecolor='white', linewidth=0.6, label='MC1R')
ax_B.bar(x + w/2, med_vals, width=w, zorder=3,
         color=[POP_COLORS[p] for p in POPS], alpha=0.30,
         edgecolor='gray', linewidth=0.6, hatch='///', label='Dataset median')
ax_B.set_xticks(x)
ax_B.set_xticklabels([POP_LABELS[p] for p in POPS],
                     fontsize=10, rotation=20, ha='right')
ax_B.set_ylabel('π (nucleotide diversity)', fontsize=11)
ax_B.set_title('MC1R π vs. dataset median across populations',
               fontsize=11.5, fontweight='bold', loc='left', pad=8)
ax_B.legend(fontsize=9.5, framealpha=0.92, edgecolor='#999', loc='upper right')
ax_B.tick_params(labelsize=10)
ax_B.spines['top'].set_visible(False)
ax_B.spines['right'].set_visible(False)
ax_B.text(-0.10, 1.10, 'B', transform=ax_B.transAxes,
          fontsize=20, fontweight='bold', va='top')

# ── Panel C ──────────────────────────────────────────────────────────────────
xmax = ymax = 1.0
ax_C.set_xlim(0, xmax)
ax_C.set_ylim(0, ymax)
ax_C.set_aspect('equal', adjustable='box')

ax_C.axhspan(thr_m, ymax, xmin=0, xmax=thr_a/xmax,
             facecolor=COLOR_C['MelanesianOnly'], alpha=0.06, zorder=0)
ax_C.axhspan(thr_m, ymax, xmin=thr_a/xmax, xmax=1,
             facecolor=COLOR_C['Both'], alpha=0.08, zorder=0)
ax_C.axhspan(0, thr_m, xmin=thr_a/xmax, xmax=1,
             facecolor=COLOR_C['AfricanOnly'], alpha=0.06, zorder=0)

ax_C.plot([0, 1], [0, 1], color='#888', lw=0.7, ls='--', zorder=1)
ax_C.axhline(thr_m, color='#888', lw=0.8, ls=':', zorder=1)
ax_C.axvline(thr_a, color='#888', lw=0.8, ls=':', zorder=1)
ax_C.text(thr_a, ymax * 0.005, f'  top 25 %\n  African',
          fontsize=7.5, color='#555', ha='left', va='bottom')
ax_C.text(xmax * 0.005, thr_m, f'  top 25 %\n  Melanesian',
          fontsize=7.5, color='#555', ha='left', va='bottom')

for q in ['Neither', 'AfricanOnly', 'MelanesianOnly', 'Both']:
    sub = df_C[df_C['quadrant'] == q]
    ax_C.scatter(sub['pbs1_african'], sub['pbs3_melanesian'],
                 c=COLOR_C[q], s=100, alpha=0.88,
                 edgecolors='white', linewidths=0.7, zorder=3,
                 label=f'{LABEL_C[q]}  (n={len(sub)})')

corner_props = dict(fontsize=9.5, fontweight='bold', alpha=0.85, zorder=2,
                    bbox=dict(boxstyle='round,pad=0.35', facecolor='white',
                              edgecolor='#999', alpha=0.92, linewidth=0.8))
ax_C.text(xmax * 0.98, ymax * 0.95, 'SHARED selection',
          color='#008572', ha='right', va='top', **corner_props)
ax_C.text(xmax * 0.02, ymax * 0.95, 'MELANESIAN-specific\nselection',
          color='#5B7A0E', ha='left', va='top', **corner_props)
ax_C.text(xmax * 0.98, thr_m * 0.92, 'AFRICAN-specific\nselection',
          color='#5B009E', ha='right', va='top', **corner_props)

texts_C = []
for _, r in df_C.iterrows():
    if r['gene'] in labels_C:
        texts_C.append(ax_C.text(r['pbs1_african'], r['pbs3_melanesian'], r['gene'],
                                  fontsize=7.5, fontweight='bold', style='italic',
                                  color='#111', zorder=5))
adjust_text(texts_C, ax=ax_C,
            arrowprops=dict(arrowstyle='-', color='#999', lw=0.5),
            force_points=(1.6, 1.6), force_text=(1.6, 1.6),
            expand_points=(2.0, 2.0), expand_text=(1.8, 1.8))

ax_C.set_xlabel('PBS on African branch  →  stronger candidate selection in Africans',
                fontsize=11, fontweight='bold', labelpad=8)
ax_C.set_ylabel('PBS on Melanesian branch  →  stronger candidate selection in Melanesians',
                fontsize=11, fontweight='bold', labelpad=8)
ax_C.set_title('Population-specific selection signals\nacross the melanogenesis network',
               fontsize=11.5, fontweight='bold', loc='left', pad=8)
ax_C.legend(loc='upper center', bbox_to_anchor=(0.5, -0.10), ncol=2,
            fontsize=8.5, framealpha=0.92, edgecolor='#999',
            title='Quadrant assignment', title_fontsize=9)
ax_C.tick_params(labelsize=10)
ax_C.text(-0.12, 1.02, 'C', transform=ax_C.transAxes,
          fontsize=20, fontweight='bold', va='top')

# ── Layout & save ─────────────────────────────────────────────────────────────
fig.subplots_adjust(left=0.07, right=0.97, top=0.96, bottom=0.14,
                    hspace=0.55, wspace=0.35)

for ext in ('png', 'pdf'):
    out = os.path.join(OUT_DIR, f'figure_composite.{ext}')
    fig.savefig(out, dpi=200, bbox_inches='tight', facecolor='white')
    print(f"saved {out}")
plt.close(fig)
