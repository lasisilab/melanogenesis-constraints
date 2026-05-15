"""
figure_gene_source_venn.py

Replacement for output/figure_gene_source_venn.png.

3-set Venn (Raghunath, GWAS Catalog, Baxter) with Raghunath at top-LEFT and
GWAS Catalog at top-RIGHT, so the upper intersection (R ∩ GWAS Catalog)
visually highlights what we consider "canonical pigmentation genes":
network-connected AND supported by independent human-genetic associations.

Bajpai 2023 CRISPR screen overlap is reported in a side panel.
"""
import os
import openpyxl
import pandas as pd
import matplotlib
matplotlib.use('Agg')
import matplotlib.pyplot as plt
from matplotlib.patches import Ellipse, FancyBboxPatch
from venny4py.venny4py import venny4py

PROJECT_DIR = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
DATA_DIR    = os.path.join(PROJECT_DIR, 'data')
OUT_DIR     = os.path.join(PROJECT_DIR, 'output')

# ── 1. Raghunath network nodes (cleaned) ───────────────────────────────────
wb = openpyxl.load_workbook(
    os.path.join(DATA_DIR, '13104_2015_1128_MOESM2_ESM.xlsx'))
ws = wb['node_properties']
NON_GENE = {
    '4HNE','ACTH','Arachidonic_Acid','cAMP','DAG','IP3','NO','PGE2','PGF2a',
    'ROS','Ca2+','PKC','Melanin','cGMP','Calcium_cyt','Ceramide','Cysteine',
    'Cysteinyl_DOPA','DHI','DHICA','DOPA','DOPAchrome','DOPAquinone',
    'Eumelanin','GSH','Glutathionyl_DOPA','Pheomelanin','Indole_5_6_quinone',
    'Indole_5_6_quinone_carboxylic_acid','Tyrosine','alpha_MSH','Nitric_oxide',
    'Singlet_oxygen','Trypsin','FICZ','Apoptosis','Cell_cycle_arrest',
    'Cell_differentiation','Cell_proliferation','Cell_survival','DNA_Damage',
    'DNA_Repair','Dendrite_formation','Lipid_Peroxidation',
    'Melanocyte_migration','Melanosome_biogenesis','Melanosome_phagocytosis',
    'Skin_aging','Skin_inflammation','UVR','UVA','UVB','Node','MMPs','PLA2',
    'PLC','PhosphodiesteHRASe',
}
raghunath = set()
for row in ws.iter_rows(min_row=2, values_only=True):
    n = str(row[0]).strip() if row[0] else ''
    if not n: continue
    for suf in ('_melan', '_kerat'):
        if n.endswith(suf): n = n[:-len(suf)]
    if ':' in n or n in NON_GENE: continue
    raghunath.add(n.upper())

# ── 2. Baxter 2018 (cross-species curated, 635 unique symbols) ─────────────
wb = openpyxl.load_workbook(
    os.path.join(DATA_DIR, 'baxter2018_650_pigmentation_genes_tableS7.xlsx'))
ws = wb['650 Pigmentation Genes']
baxter = set()
for row in ws.iter_rows(min_row=2, max_col=12, values_only=True):
    sym = row[1]
    if (sym and isinstance(sym, str)
        and not sym.startswith('ENSG') and not sym.startswith('ENSDARG')
        and len(sym) < 20):
        baxter.add(sym.strip().upper())

# ── 3. GWAS Catalog (≥2 association entries) ───────────────────────────────
gwas_df = pd.read_csv(os.path.join(DATA_DIR, 'gwas_pigmentation_associations.csv'))
gwas_df['gene'] = gwas_df['gene'].astype(str).str.upper()
gwas_counts = gwas_df.groupby('gene').size()
gwas = set(gwas_counts[gwas_counts >= 2].index)

# ── 4. Bajpai 2023 (CRISPR 10% FDR melanin-promoting) ──────────────────────
wb = openpyxl.load_workbook(
    os.path.join(DATA_DIR, 'bajpai2023_crispr_screen_tableS1.xlsx'))
ws = wb['Low SSC FACS enriched genes']
bajpai = set()
for row in ws.iter_rows(min_row=2, values_only=True):
    sym = str(row[1]).strip().upper() if row[1] else ''
    eff = row[12]; q = row[17]
    if sym and q is not None and eff is not None and q <= 0.10 and eff > 0:
        bajpai.add(sym)

print(f"sets — Raghunath:{len(raghunath)}  GWAS:{len(gwas)}  "
      f"Baxter:{len(baxter)}  Bajpai:{len(bajpai)}")

# ── 5. Compute key intersections ───────────────────────────────────────────
r_g  = raghunath & gwas
r_b  = raghunath & baxter
g_b  = gwas & baxter
r_g_b = raghunath & gwas & baxter
canon = sorted(r_g)   # "canonical pigmentation" anchor set
canon_crispr = sorted(r_g & bajpai)
print(f"R ∩ GWAS (anchor 'canonical pigmentation'): {len(canon)} genes")
print(f"  : {', '.join(canon)}")
print(f"R ∩ GWAS ∩ Baxter: {len(r_g_b)} genes")
print(f"R ∩ GWAS ∩ Bajpai: {len(canon_crispr)} genes")

# Functional-category overlap for the 8 canonical genes
import pandas as pd
cat = pd.read_csv(os.path.join(DATA_DIR, 'network_constraint_categorized.csv'))
cat['gene'] = cat['gene'].str.upper()
canon_cat = (cat[cat['gene'].isin(canon)]
             .set_index('gene')['functional_category']
             .to_dict())
pigment_specific_in_net = set(
    cat[cat['functional_category'] == 'Pigment-specific']['gene'])
canon_and_pigspec = sorted(set(canon) & pigment_specific_in_net)
pigspec_not_canon = sorted(pigment_specific_in_net - set(canon))

# ── 6. Build figure ────────────────────────────────────────────────────────
fig = plt.figure(figsize=(15, 9))
gs = fig.add_gridspec(1, 2, width_ratios=[1.35, 1.0], wspace=0.02)
ax_v   = fig.add_subplot(gs[0, 0])
ax_txt = fig.add_subplot(gs[0, 1])
ax_txt.axis('off')

# 4-set Venn. venny4py orders sets so that set 0's label sits top-FAR-LEFT
# and set 3's label sits top-FAR-RIGHT. Putting Raghunath first and GWAS
# last places them at the requested corners.
sets = {
    f'Raghunath\n({len(raghunath)})':           raghunath,
    f'Baxter\n({len(baxter)})':                 baxter,
    f'Bajpai CRISPR\n({len(bajpai)})':          bajpai,
    f'GWAS ≥2\n({len(gwas)})':                  gwas,
}
venny4py(sets=sets, asax=ax_v, colors=['#C44848', '#3D77C9', '#5DA85D',
                                       '#E8A33A'],
         line_width=1.2, font_size=11, edge_color='#333')

# venny4py text annotation positions (lifted from the library so we can
# overlay highlights). All 4 cells where Raghunath (set 0) AND GWAS (set 3)
# both appear:
#   R∩G only           → (50, 11)         n=2 (EZR, MAP3K1)
#   R∩G∩Baxter only    → (60, 17)         n=4 (KITLG, MC1R, PAX3, TYRP1)
#   R∩G∩Bajpai only    → (40, 17)         n=0 (empty cell, skip)
#   R∩G∩Baxter∩Bajpai  → (50, 35)         n=2 (OCA2, TYR)
hl_cells = [
    (50, 11, len(r_g - baxter - bajpai),
        'R ∩ GWAS only', 'EZR, MAP3K1'),
    (60, 17, len(r_g & baxter - bajpai),
        'R ∩ GWAS ∩ Baxter', 'KITLG, MC1R, PAX3, TYRP1'),
    (50, 35, len(r_g & baxter & bajpai),
        'R ∩ GWAS ∩ Baxter ∩ Bajpai', 'OCA2, TYR'),
]

# Highlight the canonical-8 cells with bold gold rings + recolor count text
for x, y, n, _, _ in hl_cells:
    if n == 0:
        continue
    ring = plt.Circle((x, y), 4.0, fill=True,
                      facecolor='#FFE066', edgecolor='#7A1A1A',
                      linewidth=2.3, zorder=2, alpha=0.95)
    ax_v.add_patch(ring)

# Bold + recolor the count text on top of each ring
for txt in list(ax_v.texts):
    try:
        px, py = txt.get_position()
    except Exception:
        continue
    for x, y, n, _, _ in hl_cells:
        if n == 0:
            continue
        if abs(px - x) < 0.5 and abs(py - y) < 0.5:
            txt.set_fontsize(13)
            txt.set_fontweight('bold')
            txt.set_color('#7A1A1A')
            txt.set_zorder(3)

ax_v.set_title(
    'Pigmentation gene sources — Raghunath ∩ GWAS (gold rings) = canonical 8',
    fontsize=13, fontweight='bold', loc='left', pad=14)

CAT_COL = {
    'Pigment-specific':  '#D94040',
    'Developmental/NC':  '#E8907E',
    'Generic signaling': '#F5C242',
    'Other':             '#B0B0B0',
}

# Header
ax_txt.text(0.0, 0.99,
            f'The {len(canon)} canonical pigmentation genes\n'
            f'(Raghunath ∩ GWAS Catalog ≥2)',
            transform=ax_txt.transAxes,
            ha='left', va='top', fontsize=13, fontweight='bold')

# Per-gene functional-category badge
y = 0.89
for g in canon:
    fc = canon_cat.get(g, '—')
    badge = CAT_COL.get(fc, '#CCCCCC')
    ax_txt.text(0.03, y, g, transform=ax_txt.transAxes,
                fontsize=12, fontweight='bold', va='center',
                family='DejaVu Sans')
    ax_txt.add_patch(plt.Rectangle((0.20, y - 0.018), 0.02, 0.022,
                                   transform=ax_txt.transAxes,
                                   facecolor=badge, edgecolor='black',
                                   linewidth=0.5))
    ax_txt.text(0.24, y, fc, transform=ax_txt.transAxes,
                fontsize=11, va='center')
    y -= 0.045

# Functional-category legend / note
y -= 0.02
ax_txt.text(0.0, y, 'Functional category (manual annotation)',
            transform=ax_txt.transAxes,
            fontsize=11, fontweight='bold', va='top')
y -= 0.04
for fc, col in CAT_COL.items():
    ax_txt.add_patch(plt.Rectangle((0.03, y - 0.018), 0.02, 0.022,
                                   transform=ax_txt.transAxes,
                                   facecolor=col, edgecolor='black',
                                   linewidth=0.5))
    ax_txt.text(0.07, y, fc, transform=ax_txt.transAxes,
                fontsize=10, va='center')
    y -= 0.038

# Overlap-with-Pigment-specific summary
y -= 0.02
ax_txt.text(0.0, y,
            f'Overlap with "Pigment-specific" category:\n'
            f'  4 of 8 canonical genes are Pigment-specific\n'
            f'    (OCA2, TYRP1, TYR, MC1R)\n'
            f'  3 Pigment-specific genes are NOT in canonical:\n'
            f'    (DCT, PMEL, MLANA — lack ≥2 GWAS associations)',
            transform=ax_txt.transAxes,
            fontsize=10.5, va='top',
            bbox=dict(boxstyle='round,pad=0.4', facecolor='#F5F5F5',
                      edgecolor='#888', lw=0.8))

for ext in ('png', 'pdf'):
    fig.savefig(os.path.join(OUT_DIR, f'figure_gene_source_venn.{ext}'),
                dpi=200, bbox_inches='tight', facecolor='white')
plt.close(fig)
print("saved figure_gene_source_venn.png/pdf")
