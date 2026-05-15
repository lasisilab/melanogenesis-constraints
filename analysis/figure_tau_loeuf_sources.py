"""
figure_tau_loeuf_sources.py

Tissue specificity (τ) vs. LoF intolerance (LOEUF) scatter.
Each gene gets exactly one combination category based on which external
pigmentation databases it appears in; each combo has a unique color + marker.

Combinations found in the 128-gene network:
  ● grey        — not in any external source  (n=94)
  ■ blue        — Baxter 2018 only            (n=20)
  ● amber       — GWAS ≥1 only               (n=4)
  ◆ green       — Bajpai 2023 only            (n=1)
  ★ purple      — GWAS + Baxter               (n=6)
  ★ crimson     — All three                   (n=3)

Output: output/figure_tau_loeuf_sources.{png,pdf}
"""
import os
import openpyxl
import numpy as np
import pandas as pd
import matplotlib
matplotlib.use('Agg')
import matplotlib.pyplot as plt
import matplotlib.lines as mlines
from scipy import stats
from adjustText import adjust_text

PROJECT_DIR = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
DATA_DIR    = os.path.join(PROJECT_DIR, 'data')
OUT_DIR     = os.path.join(PROJECT_DIR, 'output')

# ── Main data ───────────────────────────────────────────────────────────────
master = pd.read_csv(os.path.join(DATA_DIR, 'network_constraint_categorized.csv'))
master['gene'] = master['gene'].str.upper()
df = master[['gene', 'tau', 'LOEUF']].dropna(subset=['tau', 'LOEUF']).copy()
print(f"{len(df)} genes with tau + LOEUF")

# ── Gene source sets ────────────────────────────────────────────────────────
# Baxter 2018
wb = openpyxl.load_workbook(
    os.path.join(DATA_DIR, 'baxter2018_650_pigmentation_genes_tableS7.xlsx'))
ws = wb['650 Pigmentation Genes']
baxter = set()
for row in ws.iter_rows(min_row=2, max_col=12, values_only=True):
    sym = row[1]
    if (sym and isinstance(sym, str)
            and not sym.startswith('ENSG') and not sym.startswith('ENSDARG')
            and len(sym) < 20):
        baxter.add(sym.strip().upper())

# Bajpai 2023
wb = openpyxl.load_workbook(
    os.path.join(DATA_DIR, 'bajpai2023_crispr_screen_tableS1.xlsx'))
ws = wb['Low SSC FACS enriched genes']
bajpai = set()
for row in ws.iter_rows(min_row=2, values_only=True):
    sym = str(row[1]).strip().upper() if row[1] else ''
    eff = row[12]; q = row[17]
    if sym and q is not None and eff is not None and q <= 0.10 and eff > 0:
        bajpai.add(sym)

# GWAS Catalog (≥1 association)
gwas_df = pd.read_csv(os.path.join(DATA_DIR, 'gwas_pigmentation_associations.csv'))
gwas = set(gwas_df['gene'].astype(str).str.upper().unique())

df['in_gwas']   = df['gene'].isin(gwas)
df['in_baxter'] = df['gene'].isin(baxter)
df['in_bajpai'] = df['gene'].isin(bajpai)

def assign_combo(r):
    g, b, p = r['in_gwas'], r['in_baxter'], r['in_bajpai']
    if g and b and p: return 'All three (GWAS + Baxter + Bajpai)'
    if g and b:       return 'GWAS + Baxter'
    if g and p:       return 'GWAS + Bajpai'
    if b and p:       return 'Baxter + Bajpai'
    if g:             return 'GWAS only'
    if b:             return 'Baxter only'
    if p:             return 'Bajpai only'
    return 'None'

df['combo'] = df.apply(assign_combo, axis=1)
for c, sub in df.groupby('combo'):
    print(f"  {c}: n={len(sub)}")

# ── Combination styles ───────────────────────────────────────────────────────
# Order: background first, most-specific last
COMBO_ORDER = [
    'None',
    'Baxter only',
    'GWAS only',
    'Bajpai only',
    'GWAS + Baxter',
    'All three (GWAS + Baxter + Bajpai)',
]

COMBO_STYLE = {
    'None':                            dict(color='#CFCFCF', marker='o', s=50,  alpha=0.70, zorder=2, lw=0.4),
    'Baxter only':                     dict(color='#3D77C9', marker='s', s=150, alpha=0.82, zorder=3, lw=0.6),
    'GWAS only':                       dict(color='#E8A33A', marker='o', s=160, alpha=0.88, zorder=4, lw=0.6),
    'Bajpai only':                     dict(color='#5DA85D', marker='D', s=160, alpha=0.90, zorder=4, lw=0.6),
    'GWAS + Baxter':                   dict(color='#9B59B6', marker='P', s=220, alpha=0.90, zorder=5, lw=0.6),
    'All three (GWAS + Baxter + Bajpai)': dict(color='#C0392B', marker='*', s=340, alpha=0.95, zorder=6, lw=0.6),
}

COMBO_LABEL = {
    'None':                            'Not in any external source',
    'Baxter only':                     'Baxter 2018 only',
    'GWAS only':                       'GWAS Catalog ≥1 only',
    'Bajpai only':                     'Bajpai 2023 CRISPR only',
    'GWAS + Baxter':                   'GWAS + Baxter',
    'All three (GWAS + Baxter + Bajpai)': 'GWAS + Baxter + Bajpai',
}

# ── Correlation ─────────────────────────────────────────────────────────────
rho, pval = stats.spearmanr(df['tau'], df['LOEUF'])
print(f"Spearman ρ = {rho:.3f}, p = {pval:.2e}, n = {len(df)}")

# ── Labels ───────────────────────────────────────────────────────────────────
ALWAYS_LABEL = {
    'TYR', 'TYRP1', 'DCT', 'PMEL', 'MC1R', 'OCA2', 'MLANA',
    'MITF', 'SOX10', 'PAX3', 'TFAP2A', 'KIT', 'KITLG', 'EDNRB',
    # GWAS-positive genes not in the canonical set
    'CCND1', 'EZR', 'IL6', 'MAP3K1',
    # Bajpai-only
    'SPTLC2',
}

# ── Figure ───────────────────────────────────────────────────────────────────
fig, ax = plt.subplots(figsize=(11, 8.5))

legend_handles = []
for combo in COMBO_ORDER:
    sub = df[df['combo'] == combo]
    if len(sub) == 0:
        continue
    st = COMBO_STYLE[combo]
    ax.scatter(sub['tau'], sub['LOEUF'],
               c=st['color'], marker=st['marker'], s=st['s'],
               alpha=st['alpha'], edgecolors='white', linewidths=st['lw'],
               zorder=st['zorder'])
    legend_handles.append(
        mlines.Line2D([], [],
                      color=st['color'], marker=st['marker'],
                      markersize=9, linestyle='None',
                      markeredgecolor='white', markeredgewidth=0.5,
                      label=f"{COMBO_LABEL[combo]}  (n={len(sub)})"))

# Trend line
x_line = np.linspace(df['tau'].min(), df['tau'].max(), 200)
slope, intercept, *_ = stats.linregress(df['tau'], df['LOEUF'])
ax.plot(x_line, slope * x_line + intercept,
        color='#888', lw=1.0, linestyle='--', zorder=1)

# Correlation annotation
ax.text(0.03, 0.97,
        f'Spearman ρ = {rho:.3f},  p = {pval:.2e}\nn = {len(df)} genes',
        transform=ax.transAxes, fontsize=10.5, va='top',
        bbox=dict(boxstyle='round,pad=0.4', facecolor='white',
                  edgecolor='#aaa', alpha=0.92))

# Gene labels
texts = []
for _, r in df.iterrows():
    if r['gene'] in ALWAYS_LABEL:
        texts.append(ax.text(r['tau'], r['LOEUF'], r['gene'],
                             fontsize=8.5, fontweight='bold', style='italic',
                             color='#111', zorder=7))
adjust_text(texts, ax=ax,
            arrowprops=dict(arrowstyle='-', color='#aaa', lw=0.5),
            force_points=(1.2, 1.2), force_text=(1.2, 1.2),
            expand_points=(1.8, 1.8), expand_text=(1.5, 1.5))

ax.set_xlabel('Tissue specificity (τ)\n← uniform · · · tissue-specific →',
              fontsize=13, labelpad=8)
ax.set_ylabel('LOEUF  (higher = less constrained)', fontsize=13)
ax.set_title(
    'Tissue specificity (τ) vs. LoF intolerance\n'
    'across the melanogenesis network — highlighted by database source',
    fontsize=13.5, fontweight='bold', loc='left', pad=10)

ax.legend(handles=legend_handles,
          loc='upper center', bbox_to_anchor=(0.5, -0.18), ncol=2,
          fontsize=10, framealpha=0.92, edgecolor='#999',
          title='Gene source combination', title_fontsize=10.5)
ax.tick_params(labelsize=11)
ax.spines['top'].set_visible(False)
ax.spines['right'].set_visible(False)

fig.tight_layout(rect=[0, 0.18, 1, 1])

for ext in ('png', 'pdf'):
    out = os.path.join(OUT_DIR, f'figure_tau_loeuf_sources.{ext}')
    fig.savefig(out, dpi=200, bbox_inches='tight', facecolor='white')
    print(f"saved {out}")
plt.close(fig)
